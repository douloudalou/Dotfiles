from openpyxl import load_workbook
import pandas as pd
import datetime
from rich import print

#Date
now = datetime.datetime.now()
month = now.strftime("%b")
yr = now.strftime("%Y")
time = f"{month}{yr}"

#Excel 
file = r'/home/Meee/Development/Mom/FinalData.xlsx'
book = load_workbook(file)
writer = pd.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
workbook = writer.book
worksheet = workbook.create_sheet(f"{time}")
worksheet = workbook[f"{time}"]

#Retrieve Data
tableA_code = []
tableA_total = []

data = pd.read_excel(r'/home/Meee/Development/Mom/Nets Working File Aug 23.xlsx', 'Sheet4')
for c in data:
    tableA_code.append(int(data[c][0]))
    tableA_total.append(data[c][35])

print(tableA_code)
print(tableA_total)

#Paste Data
def total_data_input(data):
    data.to_excel( writer, f"{time}", index = False, startcol = 0, startrow = 0)
    print(data)

total_data = pd.DataFrame({'Code': tableA_code, 'Total': tableA_total})
total_data_input(total_data)

#End
workbook.save(file)
workbook.close()